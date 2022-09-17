
import os
from win32com import client as wc
import glob
from shutil import copyfile
import os.path
import re
import shutil as sh
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import PDFPageAggregator
from pdfminer.layout import LAParams
from pdfminer.pdfpage import PDFTextExtractionNotAllowed, PDFPage
from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
from openpyxl import Workbook

'''
step 1:
将doc、docx格式的简历转换为 pdf 文件后复制到 pdfPath 文件夹下，
将pdf格式的简历直接复制到 pdfPath 文件夹下，
'''
# 使用 Word
# word = wc.Dispatch('Word.Application')

# 使用 WPS
word = wc.Dispatch('Kwps.Application')

print('当前工作路径：' + os.getcwd())

# 处理路径
FolderPath = os.getcwd()                    # 脚本工作路径
ResultPath = 'resumes.xlsx'
PdfResultPath = 'pdfPath'
ErrorPath = 'error'
# 删除上一次的生成结果
is_result_exsist = os.path.isfile(ResultPath)

if (is_result_exsist):
    os.remove(os.path.join(FolderPath, ResultPath))

is_pdf_exsist = os.path.exists(PdfResultPath)

if (is_pdf_exsist):
    sh.rmtree(os.path.join(FolderPath, PdfResultPath))

is_error_exist = os.path.exists(ErrorPath)

if (is_error_exist):
    sh.rmtree(os.path.join(FolderPath, ErrorPath))

SaveFolderPath = FolderPath + '\\pdfPath'   # pdf 格式简历保存路径
os.mkdir(SaveFolderPath)                    # 创建文件夹
OriginalFolderPath = os.path.join(FolderPath, 'origin')  # 源文件目录
ErrorFolderPath = os.path.join(FolderPath, ErrorPath)
os.mkdir(ErrorFolderPath)
if (not os.path.exists(OriginalFolderPath)):
    os.mkdir(OriginalFolderPath)
    print('将文件放入当前目录origin文件夹后, 按任意键继续')
    input()
WordPath = OriginalFolderPath + '/*[doc, docx]'     # 筛选出doc和docx格式的文件
PdfPath = OriginalFolderPath + '/*[pdf]'            # 筛选出pdf格式的文件

print('\n简历格式转换处理中...\n')
# 将当前目录下的 doc，docx 文件转换成 pdf 文件后，放到pdfPath文件夹
files = glob.glob(WordPath)
for file_path_word in files:
    # 获取文件名
    name = os.path.basename(file_path_word)
    names = re.findall(r'(.*?).doc', name)[0]
    print(names + '.pdf')
    doc = word.Documents.Open(file_path_word)
    doc.SaveAs(SaveFolderPath + '\\%s.pdf' % names, 17)
    doc.Close()

# 将当前目录下的 pdf 文件拷贝到 pdfPath 文件夹
files = glob.glob(PdfPath)
for file_path_pdf in files:
    name = os.path.basename(file_path_pdf)
    names = re.findall(r'(.*?).pdf', name)[0]
    print(names + '.pdf')
    copyfile(file_path_pdf, SaveFolderPath + '\\%s.pdf' % names)

word.Quit()

'''
step 2:
解析pdf文件
'''


class CPdf2TxtManager():

    def changePdfToText(self, filePath):

        getInfo = {'Phone': None, 'Email': None}

        # 以二进制读模式打开
        file = open(filePath, 'rb')
        # 用文件对象来创建一个pdf文档分析器
        praser = PDFParser(file)
        # 创建一个PDF文档对象存储文档结构,提供密码初始化，没有就不用传该参数
        doc = PDFDocument(praser, password='')
        # 检查文件是否允许文本提取
        if not doc.is_extractable:
            raise PDFTextExtractionNotAllowed
        # 创建PDf 资源管理器 来管理共享资源，#caching = False不缓存
        rsrcmgr = PDFResourceManager(caching=False)
        # 创建一个PDF设备对象
        laparams = LAParams()
        # 创建一个PDF页面聚合对象
        device = PDFPageAggregator(rsrcmgr, laparams=laparams)
        # 创建一个PDF解析器对象
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        # 获得文档的目录（纲要）,文档没有纲要会报错
        # PDF文档没有目录时会报：raise PDFNoOutlines  pdfminer.pdfdocument.PDFNoOutlines
        # print(doc.get_outlines())
        # 获取page列表
        # print(PDFPage.get_pages(doc))
        # 循环遍历列表，每次处理一个page的内容
        for page in PDFPage.create_pages(doc):
            interpreter.process_page(page)
            # 接受该页面的LTPage对象
            layout = device.get_result()
            # 这里layout是一个LTPage对象 里面存放着 这个page解析出的各种对象
            # 一般包括LTTextBox, LTFigure, LTImage, LTTextBoxHorizontal 等等
            for x in layout:
                if hasattr(x, 'get_text'):
                    fileNames = os.path.splitext(filePath)
                    results = x.get_text()
                    # print('###' + results)

                    # 匹配邮箱
                    emailRegex = re.compile(r'''(
                        [a-zA-Z0-9._%+-]+ 	# 邮箱用户名
                        @ # @ symbol
                        [a-zA-Z0-9.-]+ 		# 域名
                        (.[a-zA-Z]{2,4}) 	# 域名后缀
                        )''', re.VERBOSE)

                    matchedEmail = emailRegex.search(results)
                    if matchedEmail:
                        # print(matchedEmail.group())
                        getInfo['Email'] = matchedEmail.group()

                    # 匹配手机号
                    phoneRegex = re.compile(r'''(
                        ([1])                           # 手机号码通常以‘1’开始
                        (\d{2})                         # 紧随其后有两个数字
                        (\s|-|.|'')?					# 可能有分隔符如‘-’ ‘.’ 或空格
                        (\d{4}) 						# 四个数字
                        (\s|-|.|'')? 					# 可能有分隔符如‘-’ ‘.’ 或空格
                        (\d{4}) 						# 四个数字
                        (\s*(ext|x|ext.)\s*(\d{2,5}))?  # extension
                        )''', re.VERBOSE)
                    matchedPhone = phoneRegex.search(results)
                    if matchedPhone:
                        # print(matchedPhone.group())
                        phoneNumber = matchedPhone.group()
                        phoneNumber = phoneNumber.replace(' ', '')  # 去除空格
                        phoneNumber = phoneNumber.replace('-', '')  # 去除 '-'
                        phoneNumber = phoneNumber.replace('.', '')  # 去除 '.'
                        getInfo['Phone'] = phoneNumber
        return getInfo


'''
step 3:
保存求职者信息到Excel文件
'''

print('\n简历信息提取...')

dirs = os.listdir(SaveFolderPath)   # 搜索目录
pdf2TxtManager = CPdf2TxtManager()

wb = Workbook()  # 创建文件对象
ws = wb.active  # 获取第一个sheet

# 将数据写入到指定的单元格
ws['A1'] = '姓名'
ws['B1'] = '电话'
ws['C1'] = '邮箱'

# 提取求职者的联系方式，并写入Excel文件对象
count = 0
error_count = 0
i = 2
for file in dirs:
    retInfo = pdf2TxtManager.changePdfToText(SaveFolderPath + '\\' + file)
    name = re.findall(r'(.*?).pdf', file)[0]

    is_error = False

    if (not isinstance(retInfo['Phone'], str)):
        is_error = True
        retInfo['Phone'] = '电话读取错误 请手动检查'
    if (not isinstance(retInfo['Email'], str)):
        is_error = True
        retInfo['Email'] = '邮箱读取错误 请手动检查'

    if (is_error):
        copyfile(OriginalFolderPath + '\\%s.pdf' %
                 name, ErrorFolderPath + '\\%s.pdf' % name)
        print('读取出错！！！文件已拷贝到error目录')
        error_count = error_count + 1

    print('\n <%d> ' % (i-1) + '='*50)
    print('姓名：' + name)
    print('电话：' + retInfo['Phone'])
    print('邮箱：' + retInfo['Email'])

    ws['A'+str(i)] = name  # 写入姓名
    ws['B'+str(i)] = retInfo['Phone']  # 写入电话
    ws['C'+str(i)] = retInfo['Email']  # 写入邮箱
    i = i+1
    count = count + 1

# 保存为resumes.xlsx
wb.save(FolderPath + '\\' + 'resumes.xlsx')

print('\n提取记录已保存在' + FolderPath + '\\' + 'resumes.xlsx')
print('\n提取完成，共提取文件 ' + str(count) + ' 个，提取错误文件 ' + str(error_count) + '个')
print('\n按任意键退出。')
input()
